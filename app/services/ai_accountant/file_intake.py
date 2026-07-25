"""Smart intake for spreadsheets dropped into the AI chat.

Classifies uploaded spreadsheet attachments by CONTENT (never guesses
silently — the detection is stated to the user, who can redirect) and routes:

1. Chart-of-accounts exports (گروه/کل/معین/تفصیلی shape, ``نوع تفصيلي``
   column, code+title headers) → the existing migration importer: the same
   idempotent preview is staged and rendered in chat; Confirm hits
   ``POST /migration/import/confirm`` (Owner/Accountant — enforced there).
2. Transaction/journal sheets (date + debit/credit + titles; not the chart
   shape) → the existing transaction excel-import path: parsed rows +
   auto-suggested account mappings as a confirm card; Confirm hits
   ``POST /transactions/excel-import/confirm``. Unmapped titles are flagged,
   never dropped silently.
3. Anything else → attached as Q&A context (a text digest of the sheet) so
   the user can ask "what's the total?" without recording anything.

Reuses the migration parser (`app/services/migration_import`) and the
transaction excel parser (`app/services/excel_journal_parser`) — no new
parsers. Nothing here writes to the books; every write stays confirm-gated
behind the existing endpoints.
"""
from __future__ import annotations

import hashlib
import re
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from app.services import migration_import as mig

CHART_TIERS = ("group", "kol", "moein", "tafsili")

# A message that is ONLY filesystem path(s) — the pre-fix failure mode where
# drag-and-drop pasted "/Users/…/حساب معين.xls" as text. The assistant must
# ask for a real upload instead of hallucinating file contents. Paths may
# contain spaces ("حساب معين.xls"), so match line-wise: starts like a path,
# ends in a file extension, no other sentence content.
_PATH_LINE_RE = re.compile(
    r"^\s*(?:~?/|[A-Za-z]:\\|file://)\S[^\n]*"
    r"\.(?:xls[xm]?|csv|tsv|pdf|jpe?g|png|webp|txt|docx?|xml)\s*$",
    re.IGNORECASE,
)


def is_path_only_message(message: str) -> bool:
    lines = [ln for ln in (message or "").splitlines() if ln.strip()]
    return bool(lines) and all(_PATH_LINE_RE.match(ln) for ln in lines)


@dataclass
class IntakeResult:
    kind: str                      # chart_export | transactions | context
    detected: str                  # short English detection line (persisted)
    payload: dict[str, Any] = field(default_factory=dict)   # card data for the UI
    context_text: str = ""         # Q&A context fed to the model (kind=context)


def _read_attachment(att) -> bytes | None:
    try:
        return Path(att.file_path).read_bytes()
    except OSError:
        return None


def _try_chart(files: list[tuple[str, bytes]]) -> list[tuple[str, list[dict]]] | None:
    """Parse every file with the migration parser; chart-shaped when every
    sheet has the كد/عنوان header and classifies into a chart tier."""
    parsed: list[tuple[str, list[dict]]] = []
    for name, data in files:
        try:
            rows = mig.extract_rows(name, data)
        except Exception:
            return None
        if not rows:
            return None
        kind = mig.classify_rows(rows)
        if kind not in CHART_TIERS:
            return None
        # chart exports carry balances; a code+title sheet without any balance
        # columns is more likely a product/entity list — don't claim it
        if not any(r["balance_debit"] or r["balance_credit"] for r in rows) and kind != "tafsili":
            return None
        parsed.append((name, rows))
    return parsed or None


def _try_transactions(files: list[tuple[str, bytes]]) -> dict[str, Any] | None:
    """One journal-shaped sheet → excel-import preview data (token + vouchers
    + account mappings). Multi-file transaction drops take the first sheet and
    flag the rest."""
    from app.api.transactions import _EXCEL_UPLOAD_STORE
    from app.services.excel_journal_parser import parse_excel_journal

    name, data = files[0]
    suffix = Path(name).suffix.lower()
    if suffix not in (".xlsx", ".xls", ".csv", ".tsv"):
        return None

    token = hashlib.sha256(data).hexdigest()[:16] + "_" + name
    tmp_dir = Path(tempfile.gettempdir()) / "excel_imports"
    tmp_dir.mkdir(exist_ok=True)
    store_suffix = suffix if suffix in (".csv", ".tsv") else ".xlsx"
    tmp_path = tmp_dir / f"{token}{store_suffix}"
    try:
        tmp_path.write_bytes(data)
        result = parse_excel_journal(str(tmp_path))
    except Exception:
        tmp_path.unlink(missing_ok=True)
        return None
    if not result.vouchers:
        tmp_path.unlink(missing_ok=True)
        return None

    _EXCEL_UPLOAD_STORE[token] = str(tmp_path)
    return {"token": token, "result": result,
            "extra_files": [n for n, _ in files[1:]]}


def _sheet_digest(name: str, data: bytes, max_rows: int = 30) -> str:
    """Plain-text digest of an unrecognized sheet for Q&A context."""
    try:
        rows = _raw_rows(name, data)
    except Exception:
        return f"[Attached file {name}: could not read as a spreadsheet]"
    lines = [f"Attached spreadsheet {name!r} ({len(rows)} rows). Content:"]
    for row in rows[: max_rows + 1]:
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any(cells):
            lines.append(" | ".join(cells))
    if len(rows) > max_rows + 1:
        lines.append(f"… ({len(rows) - max_rows - 1} more rows omitted)")
    return "\n".join(lines)


def _raw_rows(name: str, data: bytes) -> list[list[Any]]:
    suffix = Path(name).suffix.lower()
    if data[:2] == b"PK" or suffix == ".xlsx":
        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            return [list(r) for r in wb.worksheets[0].iter_rows(values_only=True)]
        finally:
            wb.close()
    if b"<Workbook" in data[:4096]:
        return mig._rows_from_spreadsheetml(data)
    return mig._rows_from_csv(data)


def build_spreadsheet_intake(
    db: Session, attachments: list, *, can_migrate: bool
) -> IntakeResult | None:
    """Classify the spreadsheet attachments of a chat turn and build the
    intake card. ``can_migrate`` gates the chart route (migration:write)."""
    files: list[tuple[str, bytes]] = []
    for att in attachments:
        data = _read_attachment(att)
        if data is not None:
            files.append((att.file_name or "sheet", data))
    if not files:
        return None

    chart = _try_chart(files)
    if chart is not None:
        tiers_present = {mig.classify_rows(rows) for _n, rows in chart}
        missing = [t for t in CHART_TIERS if t not in tiers_present]
        if not can_migrate:
            return IntakeResult(
                kind="context",
                detected=("Looks like a chart-of-accounts export, but importing it needs the "
                          "Owner or Accountant role — attached for questions only."),
                context_text="\n\n".join(_sheet_digest(n, d) for n, d in files),
                payload={"denied": "migration"},
            )
        hasher = hashlib.sha256()
        for _n, data in files:
            hasher.update(data)
        token = hasher.hexdigest()[:32]
        batch, summary, already_applied = mig.stage_preview_batch(db, chart, token)
        return IntakeResult(
            kind="chart_export",
            detected=(f"Looks like a chart-of-accounts export "
                      f"({', '.join(sorted(tiers_present))}) — preview staged."),
            payload={
                "token": token,
                "batch_id": str(batch.id),
                "summary": summary,
                "missing_tiers": missing,
                "already_applied": already_applied,
                "default_opening_date": mig.default_opening_date(db).isoformat(),
            },
        )

    txn = _try_transactions(files)
    if txn is not None:
        result = txn["result"]
        from sqlalchemy import select

        from app.models.account import Account
        existing_codes = {a.code for a in db.execute(select(Account)).scalars().all()}
        accounts = []
        mappings = []
        unmapped = 0
        for a in result.unique_accounts:
            exists = bool(a.suggested_code) and a.suggested_code in existing_codes
            accounts.append({
                "title1": a.title1, "title2": a.title2, "title3": a.title3,
                "suggested_code": a.suggested_code, "exists_in_chart": exists,
            })
            if exists:
                mappings.append({"title1": a.title1, "title2": a.title2,
                                 "title3": a.title3, "account_code": a.suggested_code})
            else:
                unmapped += 1
        return IntakeResult(
            kind="transactions",
            detected=(f"Looks like a transaction list — {result.total_rows} rows, "
                      f"{result.total_vouchers} voucher(s), {unmapped} unmapped account(s)."),
            payload={
                "file_token": txn["token"],
                "jalali_year": result.jalali_year,
                "total_rows": result.total_rows,
                "total_vouchers": result.total_vouchers,
                "balanced_vouchers": sum(1 for v in result.vouchers if v.is_balanced),
                "accounts": accounts,
                "account_mappings": mappings,
                "unmapped_accounts": unmapped,
                "errors": result.errors,
                "extra_files": txn["extra_files"],
            },
        )

    return IntakeResult(
        kind="context",
        detected="Attached for questions — not recognized as a chart export or transaction list.",
        context_text="\n\n".join(_sheet_digest(n, d) for n, d in files),
    )
