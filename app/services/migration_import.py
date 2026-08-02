"""Migrate from another Iranian accounting system.

Parses the standard 4-file chart export — حساب گروه (L1) / حساب كل (L2) /
حساب معين (L3) / حساب تفصيلي (L4) — plus opening balances, and applies it in a
confirm-gated flow:

- گروه→کل→معین merge into the app's chart (Iranian codes preserved; match on
  code, re-import updates names, never duplicates).
- تفصيلي rows become entities: ``حساب بانكي`` → bank entity + GL cash account
  (account number pulled from the title); ``طرف مقابل`` → client/supplier,
  inferred from the parent معین when the export carries one, else defaulted to
  client and flagged for review.
- Opening balances post as ONE balanced journal (reference
  ``MIGRATION-OPENING``); an out-of-balance source routes the difference to a
  suspense account and is surfaced in the preview — never posted unbalanced.

Files are SpreadsheetML 2003 (.xls XML, UTF-8 BOM, Persian text) as the
primary format; ``.xlsx`` and CSV with the same columns are also accepted.
"""
from __future__ import annotations

import csv
import io
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.account import Account, AccountLevel
from app.models.entity import Entity
from app.models.migration import MigrationBatch, MigrationPendingRecord
from app.models.transaction import Transaction

_SS_NS = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}

OPENING_REFERENCE = "MIGRATION-OPENING"
SUSPENSE_CODE = "3999"
SUSPENSE_NAME = "تعدیلات افتتاحیه (مهاجرت)"

_PERSIAN_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")

# Column-header aliases, matched after Persian normalization (ي→ی, ك→ک).
_HEADER_MAP = {
    "کد": "code",
    "کد حساب": "code",
    "code": "code",
    "عنوان": "title",
    "عنوان حساب": "title",
    "شرح": "title",
    "title": "title",
    "name": "title",
    "نوع تفصیلی": "detail_type",
    "نوع": "detail_type",
    "detail type": "detail_type",
    "کد معین": "parent_code",
    "معین": "parent_code",
    "گردش بدهکار": "turnover_debit",
    "گردش بستانکار": "turnover_credit",
    "مانده بدهکار": "balance_debit",
    "مانده بستانکار": "balance_credit",
    "بدهکار": "balance_debit",
    "بستانکار": "balance_credit",
    "debit": "balance_debit",
    "credit": "balance_credit",
}

# نوع تفصيلي values (normalized).
_BANK_DETAIL_TYPES = ("حساب بانکی", "بانک")
_COUNTERPARTY_DETAIL_TYPES = ("طرف مقابل", "طرف حساب")

# Parent-معین title cues for counterparty direction.
_CLIENT_MOEIN_CUES = ("دریافتنی", "بدهکاران", "مشتری")
_SUPPLIER_MOEIN_CUES = ("پرداختنی", "بستانکاران", "تامین", "تأمین", "فروشندگان")
_EMPLOYEE_MOEIN_CUES = ("حقوق", "دستمزد", "کارکنان", "پرسنل")

# Iranian systems conventionally number personnel تفصيلي codes in the 1xxxx
# range (e.g. 10001 مدرسی…) and corporate counterparties in 2xxxx. Used only
# when the export carries no معین link — and always review-flagged.
_PERSONNEL_CODE_RE = re.compile(r"^1\d{4,}$")

# Fields an imported entity still needs before it's "complete".
REQUIRED_ENTITY_FIELDS = {
    "client": ("address", "phone"),
    "supplier": ("address", "phone"),
    "bank": ("account_number", "iban"),
    "employee": ("address", "iban"),
}

TIER_ORDER = ("group", "kol", "moein", "tafsili")


class MigrationParseError(ValueError):
    """The uploaded file could not be parsed as a chart export."""


# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------

def normalize_fa(text: str | None) -> str:
    """Unify Arabic/Persian letterforms + digits, strip ZWNJ/BOM, collapse space."""
    if text is None:
        return ""
    s = str(text).translate(_PERSIAN_DIGITS)
    s = s.replace("ي", "ی").replace("ك", "ک").replace("ى", "ی")
    s = s.replace("‌", " ").replace("﻿", "")
    return re.sub(r"\s+", " ", s).strip()


def parse_amount(raw) -> int:
    """Parse an export amount into integer Rials.

    Handles Persian digits, thousand separators, and the exporter's ``.0000``
    decimal artifact (including malformed ``12345.6789.0000`` shapes where the
    dots are separators)."""
    if raw is None:
        return 0
    s = str(raw).translate(_PERSIAN_DIGITS).strip()
    if not s:
        return 0
    s = s.replace("٬", "").replace(",", "").replace(" ", "")
    neg = s.startswith("-") or (s.startswith("(") and s.endswith(")"))
    s = s.strip("()-")
    parts = s.split(".")
    if len(parts) > 2:
        # keep the tail as decimals only when it's the .0000 artifact
        s = "".join(parts[:-1]) if set(parts[-1]) <= {"0"} else "".join(parts)
    try:
        value = int(Decimal(s).to_integral_value())
    except InvalidOperation:
        return 0
    return -value if neg else value


_ACCT_NUM_RE = re.compile(r"(\d{6,26})")


def bank_account_number(title: str) -> str | None:
    """Pull an embedded account number out of a bank تفصيلي title."""
    m = _ACCT_NUM_RE.search((title or "").translate(_PERSIAN_DIGITS))
    return m.group(1) if m else None


def strip_account_number(title: str) -> str:
    """Bank title minus the embedded account number, for a clean entity name."""
    cleaned = _ACCT_NUM_RE.sub("", (title or "").translate(_PERSIAN_DIGITS))
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -–—")
    return cleaned or (title or "").strip()


# ---------------------------------------------------------------------------
# File readers (SpreadsheetML / xlsx / CSV)
# ---------------------------------------------------------------------------

def _rows_from_spreadsheetml(data: bytes) -> list[list[str | None]]:
    text = data.decode("utf-8-sig", errors="replace")
    try:
        root = ET.fromstring(text)
    except ET.ParseError as exc:
        raise MigrationParseError(f"Invalid SpreadsheetML XML: {exc}") from exc
    out: list[list[str | None]] = []
    for row in root.findall(".//ss:Worksheet/ss:Table/ss:Row", _SS_NS):
        cells: list[str | None] = []
        for cell in row.findall("ss:Cell", _SS_NS):
            d = cell.find("ss:Data", _SS_NS)
            cells.append(d.text if d is not None else None)
        out.append(cells)
    return out


def _rows_from_xlsx(data: bytes) -> list[list[str | None]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        return [
            [(str(c) if c is not None else None) for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
    finally:
        wb.close()


def _rows_from_csv(data: bytes) -> list[list[str | None]]:
    text = data.decode("utf-8-sig", errors="replace")
    return [list(r) for r in csv.reader(io.StringIO(text))]


def extract_rows(filename: str, data: bytes) -> list[dict]:
    """Read one export file into normalized row dicts (any supported format)."""
    if data[:2] == b"PK":
        raw_rows = _rows_from_xlsx(data)
    elif data.lstrip(b"\xef\xbb\xbf").lstrip()[:5] == b"<?xml" or b"<Workbook" in data[:4096]:
        raw_rows = _rows_from_spreadsheetml(data)
    else:
        raw_rows = _rows_from_csv(data)

    header_idx, mapping = None, {}
    for i, row in enumerate(raw_rows[:10]):
        norm = [normalize_fa(c).lower() if c else "" for c in row]
        cols = {j: _HEADER_MAP[h] for j, h in enumerate(norm) if h in _HEADER_MAP}
        if "code" in cols.values() and "title" in cols.values():
            header_idx, mapping = i, cols
            break
    if header_idx is None:
        raise MigrationParseError(f"{filename}: header row (كد / عنوان) not found")

    rows: list[dict] = []
    for raw in raw_rows[header_idx + 1:]:
        vals = {key: (raw[j] if j < len(raw) else None) for j, key in mapping.items()}
        code = normalize_fa(vals.get("code"))
        if not code:  # the totals row has no code
            continue
        rows.append({
            "code": code,
            "title": normalize_fa(vals.get("title")),
            "detail_type": normalize_fa(vals.get("detail_type")) or None,
            "parent_code": normalize_fa(vals.get("parent_code")) or None,
            "balance_debit": parse_amount(vals.get("balance_debit")),
            "balance_credit": parse_amount(vals.get("balance_credit")),
        })
    return rows


def classify_rows(rows: list[dict]) -> str:
    """Which tier a parsed file holds: group / kol / moein / tafsili."""
    if any(r["detail_type"] for r in rows):
        return "tafsili"
    lengths = {len(r["code"]) for r in rows if r["code"].isdigit()}
    if lengths and max(lengths) <= 2:
        return "group"
    if lengths and max(lengths) <= 4:
        return "kol"
    if lengths and max(lengths) <= 6:
        return "moein"
    return "tafsili"


def is_bank_row(row: dict) -> bool:
    dt = row.get("detail_type") or ""
    return any(t in dt for t in _BANK_DETAIL_TYPES)


def is_counterparty_row(row: dict) -> bool:
    dt = row.get("detail_type") or ""
    if any(t in dt for t in _COUNTERPARTY_DETAIL_TYPES):
        return True
    return bool(dt) and not is_bank_row(row)


# ---------------------------------------------------------------------------
# Preview
# ---------------------------------------------------------------------------

def infer_counterparty_type(row: dict, moein_by_code: dict[str, dict]) -> tuple[str, bool]:
    """(entity_type, ambiguous?) for a طرف مقابل row.

    Uses the parent معین title when the export links one (``کد معین`` column,
    or a tafsili code prefixed with a known معین code). Without a link, the
    personnel code convention (1xxxx = پرسنل) marks employees; everything else
    defaults to client — both review-flagged."""
    parent = None
    pc = row.get("parent_code")
    if pc and pc in moein_by_code:
        parent = moein_by_code[pc]
    else:
        for length in (6, 4):
            prefix = row["code"][:length]
            if len(row["code"]) > length and prefix in moein_by_code:
                parent = moein_by_code[prefix]
                break
    if parent:
        title = parent["title"]
        if any(cue in title for cue in _EMPLOYEE_MOEIN_CUES):
            return "employee", False
        if any(cue in title for cue in _CLIENT_MOEIN_CUES):
            return "client", False
        if any(cue in title for cue in _SUPPLIER_MOEIN_CUES):
            return "supplier", False
    if _PERSONNEL_CODE_RE.match(row["code"]):
        return "employee", True
    return "client", True


def build_preview(parsed_files: list[tuple[str, list[dict]]]) -> tuple[dict, dict]:
    """(payload, summary) for a set of parsed files.

    ``payload`` is the tier→rows dict staged on the batch; ``summary`` is the
    user-facing preview: tier counts, تفصیلی split, opening totals + balance
    check, and a validation report."""
    payload: dict[str, list[dict]] = {}
    errors: list[str] = []
    warnings: list[str] = []
    files_meta: list[dict] = []

    for filename, rows in parsed_files:
        kind = classify_rows(rows)
        files_meta.append({"filename": filename, "kind": kind, "rows": len(rows)})
        if kind in payload:
            errors.append(f"Duplicate tier '{kind}': {filename}")
        else:
            payload[kind] = rows

    if not payload:
        errors.append("No recognizable chart files uploaded")

    # Duplicate codes within a tier
    for kind, rows in payload.items():
        seen: set[str] = set()
        for r in rows:
            if r["code"] in seen:
                errors.append(f"Duplicate code {r['code']} in {kind} file")
            seen.add(r["code"])

    groups = {r["code"] for r in payload.get("group", [])}
    kols = {r["code"] for r in payload.get("kol", [])}
    unmatched: list[str] = []
    if groups:
        for r in payload.get("kol", []):
            if r["code"][:2] not in groups:
                unmatched.append(f"کل {r['code']} ({r['title']}) has no parent گروه")
    if kols:
        for r in payload.get("moein", []):
            if r["code"][:4] not in kols:
                unmatched.append(f"معین {r['code']} ({r['title']}) has no parent کل")
    warnings.extend(unmatched)

    # Opening balances from the finest tier available
    basis = next((t for t in ("moein", "kol", "group") if t in payload), None)
    total_debit = sum(r["balance_debit"] for r in payload.get(basis, [])) if basis else 0
    total_credit = sum(r["balance_credit"] for r in payload.get(basis, [])) if basis else 0
    difference = total_debit - total_credit

    tafsili = payload.get("tafsili", [])
    moein_by_code = {r["code"]: r for r in payload.get("moein", [])}
    banks = [r for r in tafsili if is_bank_row(r)]
    counterparties = [r for r in tafsili if not is_bank_row(r) and is_counterparty_row(r)]
    ambiguous_count = 0
    cp_preview = []
    for r in counterparties:
        etype, ambiguous = infer_counterparty_type(r, moein_by_code)
        ambiguous_count += 1 if ambiguous else 0
        cp_preview.append({
            "code": r["code"], "title": r["title"], "entity_type": etype,
            "ambiguous": ambiguous,
            "balance_debit": r["balance_debit"], "balance_credit": r["balance_credit"],
        })
    bank_preview = [{
        "code": r["code"], "title": r["title"],
        "name": strip_account_number(r["title"]),
        "account_number": bank_account_number(r["title"]),
        "balance_debit": r["balance_debit"], "balance_credit": r["balance_credit"],
    } for r in banks]

    if ambiguous_count:
        warnings.append(
            f"{ambiguous_count} counterparties have no معین link — defaulting to client, flagged for review"
        )
    if difference != 0:
        warnings.append(
            f"Opening balances are out of balance by {abs(difference)} — the difference will post to the "
            f"suspense account {SUSPENSE_CODE} ({SUSPENSE_NAME})"
        )

    summary = {
        "files": files_meta,
        "tiers": {t: len(payload.get(t, [])) for t in TIER_ORDER if t in payload},
        "tafsili_split": {"bank_accounts": len(banks), "counterparties": len(counterparties)},
        "counterparty_types": {
            t: sum(1 for c in cp_preview if c["entity_type"] == t)
            for t in sorted({c["entity_type"] for c in cp_preview})
        },
        "banks": bank_preview,
        "counterparties": cp_preview,
        "opening": {
            "basis": basis,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "balanced": difference == 0,
            "difference": difference,
            "suspense_needed": difference != 0,
        },
        "validation": {"errors": errors, "warnings": warnings},
    }
    return payload, summary


def stage_preview_batch(
    db: Session, parsed_files: list[tuple[str, list[dict]]], token: str
) -> tuple[MigrationBatch, dict, bool]:
    """Build the preview and upsert the staged ``MigrationBatch`` for it.

    Shared by the migration page's preview endpoint and the chat smart-intake
    (dropping chart exports into the chat). Returns
    (batch, summary, already_applied). Flushes; the caller commits + audits.
    """
    payload, summary = build_preview(parsed_files)

    already_applied = False
    batch = db.execute(
        select(MigrationBatch).where(MigrationBatch.token == token)
    ).scalars().first()
    if batch is not None and batch.status == "applied":
        already_applied = True
        summary["validation"]["warnings"].append(
            "These exact files were already imported — confirming again re-applies as an update (no duplicates)."
        )
        batch.status = "pending"
        batch.payload = payload
        batch.summary = summary
    elif batch is not None:
        batch.payload = payload
        batch.summary = summary
        batch.source_files = summary["files"]
    else:
        batch = MigrationBatch(
            token=token, status="pending", payload=payload,
            summary=summary, source_files=summary["files"],
        )
        db.add(batch)
    db.flush()
    return batch, summary, already_applied


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------

_TIER_LEVEL = {"group": AccountLevel.GROUP, "kol": AccountLevel.GENERAL, "moein": AccountLevel.SUB}


def _merge_chart(db: Session, payload: dict) -> dict:
    """Create/merge گروه→کل→معین by code. Idempotent: existing codes update
    their name, never duplicate. Returns per-tier created/updated counts."""
    stats = {t: {"created": 0, "updated": 0} for t in ("group", "kol", "moein")}
    by_code: dict[str, Account] = {}

    def _get(code: str) -> Account | None:
        if code in by_code:
            return by_code[code]
        acc = db.execute(select(Account).where(Account.code == code)).scalars().first()
        if acc:
            by_code[code] = acc
        return acc

    # Parent prefix length per tier, derived from the actual codes uploaded.
    tier_codes = {t: [r["code"] for r in payload.get(t, [])] for t in ("group", "kol", "moein")}
    parent_tier = {"kol": "group", "moein": "kol"}

    for tier in ("group", "kol", "moein"):
        for row in payload.get(tier, []):
            acc = _get(row["code"])
            if acc is not None:
                if row["title"] and normalize_fa(acc.name) != row["title"]:
                    acc.name = row["title"]
                    stats[tier]["updated"] += 1
                continue
            parent = None
            pt = parent_tier.get(tier)
            if pt:
                plen = max((len(c) for c in tier_codes.get(pt, [])), default=0)
                if plen and len(row["code"]) > plen:
                    parent = _get(row["code"][:plen])
            acc = Account(
                code=row["code"], name=row["title"],
                level=_TIER_LEVEL[tier], parent_id=(parent.id if parent else None),
            )
            db.add(acc)
            db.flush()
            by_code[row["code"]] = acc
            stats[tier]["created"] += 1
    return stats


def _find_imported_entity(db: Session, etype: str, name: str, account_number: str | None) -> Entity | None:
    if account_number:
        found = db.execute(
            select(Entity).where(Entity.type == etype, Entity.account_number == account_number)
        ).scalars().first()
        if found:
            return found
    return db.execute(
        select(Entity).where(Entity.type == etype, Entity.name.ilike(name))
    ).scalars().first()


def _import_entities(db: Session, payload: dict, locale: str) -> tuple[dict, list[dict]]:
    """تفصيلي rows → entities. Banks get (or reuse) a GL cash account via the
    same auto-link used everywhere else. Idempotent by type+name (banks also by
    account number). Returns (stats, entity_records) where entity_records feed
    the completion queue."""
    from app.services.ai_accountant.entity_create import _apply_details, _resolve_bank_account

    stats = {"banks_created": 0, "banks_reused": 0,
             "counterparties_created": 0, "counterparties_reused": 0}
    records: list[dict] = []
    moein_by_code = {r["code"]: r for r in payload.get("moein", [])}

    for row in payload.get("tafsili", []):
        if is_bank_row(row):
            acct_no = bank_account_number(row["title"])
            name = strip_account_number(row["title"])
            entity = _find_imported_entity(db, "bank", name, acct_no)
            if entity is None:
                entity = Entity(type="bank", name=name)
                db.add(entity)
                code, _created = _resolve_bank_account(db, name, None, locale)
                entity.code = code
                stats["banks_created"] += 1
            else:
                if not entity.code:
                    code, _created = _resolve_bank_account(db, name, None, locale)
                    entity.code = code
                stats["banks_reused"] += 1
            _apply_details(entity, {"account_number": acct_no, "bank_name": name}, only_blank=True)
            db.flush()
            records.append({
                "entity": entity, "source_code": row["code"], "flags": [],
                "balance": row["balance_debit"] - row["balance_credit"],
                "is_bank": True,
            })
        elif is_counterparty_row(row):
            etype, ambiguous = infer_counterparty_type(row, moein_by_code)
            name = row["title"]
            entity = _find_imported_entity(db, etype, name, None)
            if entity is None:
                # A re-import may find it under a different inferred type
                # (e.g. previously defaulted to client, now recognized as an
                # employee) — reuse and retype instead of duplicating.
                for other in ("client", "supplier", "employee"):
                    if other == etype:
                        continue
                    entity = _find_imported_entity(db, other, name, None)
                    if entity is not None:
                        entity.type = etype
                        break
            if entity is None:
                entity = Entity(type=etype, name=name)
                db.add(entity)
                db.flush()
                stats["counterparties_created"] += 1
            else:
                stats["counterparties_reused"] += 1
            records.append({
                "entity": entity, "source_code": row["code"],
                "flags": (["type_ambiguous"] if ambiguous else []),
                "balance": row["balance_debit"] - row["balance_credit"],
                "is_bank": False,
            })
    return stats, records


def missing_entity_fields(entity: Entity) -> list[str]:
    required = REQUIRED_ENTITY_FIELDS.get(entity.type, ())
    return [f for f in required if not (getattr(entity, f, None) or "").strip()]


def _queue_pending_records(db: Session, batch: MigrationBatch, records: list[dict]) -> int:
    """Create/update "Complete imported records" queue rows. Non-blocking."""
    count = 0
    for rec in records:
        entity: Entity = rec["entity"]
        missing = missing_entity_fields(entity)
        flags = rec["flags"]
        existing = db.execute(
            select(MigrationPendingRecord).where(
                MigrationPendingRecord.entity_id == entity.id,
                MigrationPendingRecord.status == "pending",
            )
        ).scalars().first()
        if not missing and not flags:
            if existing:
                existing.status = "resolved"
                existing.resolved_at = datetime.now(timezone.utc)
            continue
        if existing:
            existing.batch_id = batch.id
            existing.missing_fields = missing
            existing.review_flags = flags
            existing.entity_type = entity.type
            existing.source_code = rec["source_code"]
        else:
            db.add(MigrationPendingRecord(
                batch_id=batch.id, entity_id=entity.id, entity_type=entity.type,
                source_code=rec["source_code"], missing_fields=missing,
                review_flags=flags, status="pending",
            ))
        count += 1
    db.flush()
    return count


def _bank_cash_moein(payload: dict, bank_total: int) -> dict | None:
    """The معین that aggregates the bank تفصيلي balances (e.g. موجودی
    بانک‌های ریالی), so the opening journal can split it into the per-bank GL
    accounts. Only splits when the معین's debit balance covers the bank sum."""
    candidates = [
        r for r in payload.get("moein", [])
        if "بانک" in r["title"] and (r["balance_debit"] - r["balance_credit"]) >= bank_total
    ]
    if not candidates:
        return None
    # prefer the tightest fit (usually the exact aggregate account)
    return min(candidates, key=lambda r: (r["balance_debit"] - r["balance_credit"]) - bank_total)


def _build_opening_lines(db: Session, payload: dict, bank_records: list[dict]) -> tuple[list[dict], dict]:
    """Journal lines from the basis tier, with the bank معین split into the
    per-bank GL cash accounts created for the تفصيلي banks."""
    basis = next((t for t in ("moein", "kol", "group") if t in payload), None)
    if basis is None:
        return [], {"basis": None}

    rows = [dict(r) for r in payload[basis]]
    info: dict = {"basis": basis, "bank_split": False}

    bank_lines: list[dict] = []
    if basis == "moein" and bank_records:
        bank_total = sum(r["balance"] for r in bank_records)
        target = _bank_cash_moein(payload, bank_total)
        if target is not None and bank_total > 0:
            for r in bank_records:
                if r["balance"] == 0:
                    continue
                entity: Entity = r["entity"]
                bank_lines.append({
                    "account_code": entity.code,
                    "debit": max(r["balance"], 0),
                    "credit": max(-r["balance"], 0),
                    "line_description": entity.name,
                })
            for row in rows:
                if row["code"] == target["code"]:
                    row["balance_debit"] -= bank_total
                    break
            info["bank_split"] = True
            info["bank_split_moein"] = target["code"]

    lines: list[dict] = []
    for row in rows:
        net = row["balance_debit"] - row["balance_credit"]
        if net == 0:
            continue
        lines.append({
            "account_code": row["code"],
            "debit": max(net, 0),
            "credit": max(-net, 0),
            "line_description": row["title"],
        })
    lines.extend(bank_lines)

    total_debit = sum(l["debit"] for l in lines)
    total_credit = sum(l["credit"] for l in lines)
    diff = total_debit - total_credit
    if diff != 0:
        lines.append({
            "account_code": SUSPENSE_CODE,
            "debit": max(-diff, 0),
            "credit": max(diff, 0),
            "line_description": SUSPENSE_NAME,
        })
        info["suspense_amount"] = abs(diff)
    info["total_debit"] = max(total_debit, total_credit)
    return lines, info


def _post_opening_journal(db: Session, lines: list[dict], opening_date: date, locale: str) -> tuple[Transaction | None, bool]:
    """Post the opening journal, replacing a previous migration opening journal
    if one exists (re-import updates, never duplicates). Returns
    (transaction, replaced_previous)."""
    from app.api.transactions import _create_transaction_from_payload
    from app.schemas.transaction import TransactionCreate
    from app.services.account_resolver import _ensure_account

    if not lines:
        return None, False

    if any(l["account_code"] == SUSPENSE_CODE for l in lines):
        _ensure_account(db, SUSPENSE_CODE, SUSPENSE_NAME, locale)

    replaced = False
    previous = db.execute(
        select(Transaction).where(
            Transaction.reference == OPENING_REFERENCE,
            Transaction.deleted_at.is_(None),
        )
    ).scalars().all()
    for prev in previous:
        prev.deleted_at = datetime.now(timezone.utc)
        replaced = True
    if previous:
        db.flush()

    payload = TransactionCreate(
        date=opening_date,
        reference=OPENING_REFERENCE,
        description="تراز افتتاحیه — مهاجرت از سیستم قبلی / Opening balances migrated from previous system",
        currency="IRR",
        lines=lines,
    )
    txn = _create_transaction_from_payload(db, payload)
    return txn, replaced


def default_opening_date(db: Session) -> date:
    """Start of the current fiscal period: Farvardin 1 when the display
    calendar is Jalali, else January 1 — nudged past a closed period."""
    from app.services.locale_service import get_display_calendar
    from app.services.period_service import get_closed_period
    from app.utils.jalali import gregorian_to_jalali, jalali_to_gregorian

    today = date.today()
    if get_display_calendar(db) == "jalali":
        jy, _m, _d = gregorian_to_jalali(today)
        start = jalali_to_gregorian(jy, 1, 1)
    else:
        start = date(today.year, 1, 1)
    closed = get_closed_period(db)
    if closed and closed >= start:
        from datetime import timedelta
        start = closed + timedelta(days=1)
    return min(start, today)


def apply_batch(db: Session, batch: MigrationBatch, opening_date: date) -> dict:
    """Apply a previewed batch: chart merge → entities → opening journal →
    completion queue. Flushes; the caller commits + audits."""
    from app.services.locale_service import get_reporting_locale

    locale = (get_reporting_locale(db) or "default").strip().lower()
    payload = batch.payload or {}

    chart_stats = _merge_chart(db, payload)
    entity_stats, records = _import_entities(db, payload, locale)
    bank_records = [r for r in records if r["is_bank"]]
    lines, journal_info = _build_opening_lines(db, payload, bank_records)
    txn, replaced = _post_opening_journal(db, lines, opening_date, locale)
    if txn is not None:
        # Link every migrated party with an opening balance to the journal, so
        # entity-level activity, AR/AP and "who are our clients" views see the
        # migrated books — not just the aggregate معین totals.
        from app.models.entity import TransactionEntity

        linked = 0
        for rec in records:
            if rec["balance"] == 0:
                continue
            db.add(TransactionEntity(
                transaction_id=txn.id, entity_id=rec["entity"].id,
                role=rec["entity"].type,
                amount=rec["balance"],  # the entity's own opening balance
            ))
            linked += 1
        db.flush()
        journal_info["entity_links"] = linked
    pending_count = _queue_pending_records(db, batch, records)

    batch.status = "applied"
    batch.opening_date = opening_date
    batch.opening_transaction_id = txn.id if txn is not None else None
    batch.applied_at = datetime.now(timezone.utc)
    result = {
        "chart": chart_stats,
        "entities": entity_stats,
        "opening_journal": {
            **journal_info,
            "transaction_id": str(txn.id) if txn is not None else None,
            "opening_date": opening_date.isoformat(),
            "lines": len(lines),
            "replaced_previous": replaced,
        },
        "pending_records": pending_count,
    }
    batch.result = result
    db.flush()
    return result
